"""
FinancialStatements - Báo cáo tài chính theo Thông tư 99/2025/TT-BTC.

Module này cung cấp:
- Bảng tình hình tài chính (thay thế Bảng cân đối kế toán)
- Báo cáo kết quả hoạt động kinh doanh
- Xuất báo cáo ra Excel

Tuân thủ Thông tư 99/2025/TT-BTC

Example:
    >>> from reports.financial_statements import FinancialStatements
    >>> fs = FinancialStatements()
    >>> report = fs.statement_of_financial_position(to_date="2025-12-31")
"""

# reports/financial_statements.py
# BÁO CÁO TÀI CHÍNH THEO THÔNG TƯ 99/2025/TT-BTC
# CODE MỚI - KHÔNG ẢNH HƯỞNG CODE CŨ

import sqlite3
import os
import sys
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# Thêm đường dẫn gốc để import module cũ (nếu cần, nhưng không bắt buộc)
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

class FinancialStatements:
    """
    Báo cáo tài chính theo Thông tư 99/2025/TT-BTC
    Chỉ THÊM MỚI, không sửa code cũ
    """
    
    def __init__(self, db_path=None):
        # Tìm đường dẫn database (giữ nguyên logic cũ)
        if db_path is None:
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            possible_paths = [
                os.path.join(base_dir, "ke_toan_data", "ke_toan.db"),
                os.path.join(base_dir, "data", "ke_toan.db"),
                os.path.join(base_dir, "dist", "ke_toan_data", "ke_toan.db"),
                "ke_toan_data/ke_toan.db"
            ]
            for path in possible_paths:
                if os.path.exists(path):
                    db_path = path
                    break
        
        self.db_path = db_path
        print(f"[FinancialStatements] Kết nối DB: {self.db_path}")
    
    def _get_connection(self):
        """Lấy kết nối database"""
        return sqlite3.connect(self.db_path)
    
    def _get_account_balance(self, account_code, to_date):
        """
        Lấy số dư tài khoản đến ngày chỉ định
        Dùng chung với logic cũ (bảng journal_entries)
        """
        conn = self._get_connection()
        cursor = conn.cursor()
        
        # Dựa vào cấu trúc bảng journal_entries hiện có
        try:
            cursor.execute("""
                SELECT COALESCE(SUM(debit - credit), 0)
                FROM journal_entries 
                WHERE account_code = ? AND date <= ?
            """, (account_code, to_date))
            result = cursor.fetchone()[0]
        except:
            # Thử với tên cột khác nếu cấu trúc khác
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0)
                FROM journal_entries 
                WHERE account_code = ? AND date <= ? AND entry_type = 'debit'
            """, (account_code, to_date))
            debit = cursor.fetchone()[0]
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0)
                FROM journal_entries 
                WHERE account_code = ? AND date <= ? AND entry_type = 'credit'
            """, (account_code, to_date))
            credit = cursor.fetchone()[0]
            result = debit - credit
        
        conn.close()
        return result
    
    def _get_accounts_by_type(self, account_type):
        """
        Lấy danh sách tài khoản theo loại
        Dựa vào bảng accounts hiện có
        """
        conn = self._get_connection()
        cursor = conn.cursor()
        
        # Kiểm tra cấu trúc bảng accounts
        cursor.execute("PRAGMA table_info(accounts)")
        columns = [col[1] for col in cursor.fetchall()]
        
        if 'type' in columns:
            cursor.execute("SELECT code, name FROM accounts WHERE type = ?", (account_type,))
        elif 'account_type' in columns:
            cursor.execute("SELECT code, name FROM accounts WHERE account_type = ?", (account_type,))
        else:
            # Fallback: lấy tất cả
            cursor.execute("SELECT code, name FROM accounts")
        
        results = cursor.fetchall()
        conn.close()
        return results
    
    def statement_of_financial_position(self, to_date=None):
        """
        BẢNG TÌNH HÌNH TÀI CHÍNH
        (thay thế Bảng cân đối kế toán - theo Thông tư 99)
        """
        if to_date is None:
            to_date = datetime.now().strftime("%Y-%m-%d")
        
        # Định nghĩa nhóm tài khoản theo Thông tư 99
        account_groups = {
            "TÀI SẢN NGẮN HẠN": {
                "codes": ["111", "112", "113", "131", "136", "138", "141", "152", "153", "154", "155", "156", "157", "158"],
                "accounts": []
            },
            "TÀI SẢN DÀI HẠN": {
                "codes": ["211", "212", "213", "214", "215", "217", "221", "222", "228", "229", "242", "244"],
                "accounts": []
            },
            "NỢ PHẢI TRẢ": {
                "codes": ["331", "333", "334", "335", "336", "337", "338", "341", "342", "343", "344", "347", "352", "356"],
                "accounts": []
            },
            "VỐN CHỦ SỞ HỮU": {
                "codes": ["411", "412", "413", "414", "415", "417", "418", "419", "421", "431"],
                "accounts": []
            }
        }
        
        # Lấy thông tin tài khoản từ database
        all_accounts = {}
        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT code, name FROM accounts")
        for row in cursor.fetchall():
            all_accounts[row[0]] = row[1]
        conn.close()
        
        # Tính số dư cho từng nhóm
        for group_name, group_info in account_groups.items():
            total = 0
            for code_pattern in group_info["codes"]:
                # Tìm tài khoản theo pattern
                for acc_code, acc_name in all_accounts.items():
                    if acc_code.startswith(code_pattern) or acc_code == code_pattern:
                        balance = self._get_account_balance(acc_code, to_date)
                        total += balance
                        group_info["accounts"].append({
                            "code": acc_code,
                            "name": acc_name,
                            "balance": balance
                        })
            group_info["total"] = total
        
        total_assets = account_groups["TÀI SẢN NGẮN HẠN"]["total"] + account_groups["TÀI SẢN DÀI HẠN"]["total"]
        total_liabilities_equity = account_groups["NỢ PHẢI TRẢ"]["total"] + account_groups["VỐN CHỦ SỞ HỮU"]["total"]
        
        return {
            "report_name": "BẢNG TÌNH HÌNH TÀI CHÍNH",
            "report_date": to_date,
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "groups": account_groups,
            "total_assets": total_assets,
            "total_liabilities_equity": total_liabilities_equity,
            "difference": total_assets - total_liabilities_equity
        }
    
    def income_statement(self, from_date, to_date):
        """
        BÁO CÁO KẾT QUẢ HOẠT ĐỘNG KINH DOANH
        """
        conn = self._get_connection()
        cursor = conn.cursor()
        
        # Lấy tất cả tài khoản doanh thu (5xx, 6xx)
        cursor.execute("SELECT code, name FROM accounts WHERE code LIKE '5%' OR code LIKE '6%'")
        revenue_accounts = cursor.fetchall()
        
        # Lấy tất cả tài khoản chi phí (6xx, 7xx, 8xx)
        cursor.execute("SELECT code, name FROM accounts WHERE code LIKE '6%' OR code LIKE '7%' OR code LIKE '8%'")
        expense_accounts = cursor.fetchall()
        conn.close()
        
        total_revenue = 0
        revenue_details = []
        
        for acc_code, acc_name in revenue_accounts:
            # Doanh thu: số dư có (credit)
            conn = self._get_connection()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT COALESCE(SUM(credit - debit), 0)
                FROM journal_entries 
                WHERE account_code = ? AND date BETWEEN ? AND ?
            """, (acc_code, from_date, to_date))
            amount = cursor.fetchone()[0]
            conn.close()
            
            if amount != 0:
                total_revenue += amount
                revenue_details.append({"code": acc_code, "name": acc_name, "amount": amount})
        
        total_expense = 0
        expense_details = []
        
        conn = self._get_connection()
        cursor = conn.cursor()
        for acc_code, acc_name in expense_accounts:
            cursor.execute("""
                SELECT COALESCE(SUM(debit - credit), 0)
                FROM journal_entries 
                WHERE account_code = ? AND date BETWEEN ? AND ?
            """, (acc_code, from_date, to_date))
            amount = cursor.fetchone()[0]
            
            if amount != 0:
                total_expense += amount
                expense_details.append({"code": acc_code, "name": acc_name, "amount": amount})
        conn.close()
        
        net_profit = total_revenue - total_expense
        
        return {
            "report_name": "BÁO CÁO KẾT QUẢ HOẠT ĐỘNG KINH DOANH",
            "from_date": from_date,
            "to_date": to_date,
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "revenue": {
                "total": total_revenue,
                "details": revenue_details
            },
            "expense": {
                "total": total_expense,
                "details": expense_details
            },
            "net_profit": net_profit
        }
    
    def export_to_excel(self, report_type, params, output_path=None):
        """
        Xuất báo cáo ra Excel
        KHÔNG hardcode font - để Excel dùng font mặc định hệ thống
        """
        if output_path is None:
            output_dir = os.path.join(os.path.dirname(self.db_path), "exports") if self.db_path else "exports"
            os.makedirs(output_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(output_dir, f"{report_type}_{timestamp}.xlsx")
        
        if report_type == "financial_position":
            data = self.statement_of_financial_position(**params)
            return self._export_balance_sheet(data, output_path)
        elif report_type == "income":
            data = self.income_statement(**params)
            return self._export_income_statement(data, output_path)
        else:
            raise ValueError(f"Unknown report type: {report_type}")
    
    def _export_balance_sheet(self, data, output_path):
        """Xuất Bảng tình hình tài chính - KHÔNG hardcode font"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Bảng tình hình tài chính"
        
        # Header - KHÔNG hardcode font, để mặc định
        ws['A1'] = data['report_name']
        ws['A2'] = f"Ngày {data['report_date']}"
        ws['A3'] = f"Ngày xuất: {data['created_at']}"
        
        row = 5
        for group_name, group_info in data['groups'].items():
            ws[f'A{row}'] = group_name
            ws[f'B{row}'] = "Mã TK"
            ws[f'C{row}'] = "Tên tài khoản"
            ws[f'D{row}'] = "Số dư"
            row += 1
            
            for acc in group_info['accounts']:
                ws[f'A{row}'] = ""
                ws[f'B{row}'] = acc['code']
                ws[f'C{row}'] = acc['name']
                ws[f'D{row}'] = acc['balance']
                row += 1
            
            ws[f'B{row}'] = "CỘNG"
            ws[f'D{row}'] = group_info['total']
            row += 2
        
        row += 1
        ws[f'B{row}'] = "TỔNG CỘNG TÀI SẢN"
        ws[f'D{row}'] = data['total_assets']
        row += 1
        ws[f'B{row}'] = "TỔNG CỘNG NGUỒN VỐN"
        ws[f'D{row}'] = data['total_liabilities_equity']
        
        # Tự động điều chỉnh độ rộng cột
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20
        
        wb.save(output_path)
        return output_path
    
    def _export_income_statement(self, data, output_path):
        """Xuất Báo cáo KQKD - KHÔNG hardcode font"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Báo cáo KQKD"
        
        ws['A1'] = data['report_name']
        ws['A2'] = f"Từ {data['from_date']} đến {data['to_date']}"
        ws['A3'] = f"Ngày xuất: {data['created_at']}"
        
        row = 5
        ws[f'A{row}'] = "I. DOANH THU"
        ws[f'B{row}'] = "Mã TK"
        ws[f'C{row}'] = "Diễn giải"
        ws[f'D{row}'] = "Số tiền"
        row += 1
        
        for item in data['revenue']['details']:
            ws[f'A{row}'] = ""
            ws[f'B{row}'] = item['code']
            ws[f'C{row}'] = item['name']
            ws[f'D{row}'] = item['amount']
            row += 1
        
        ws[f'C{row}'] = "Tổng doanh thu"
        ws[f'D{row}'] = data['revenue']['total']
        row += 2
        
        ws[f'A{row}'] = "II. CHI PHÍ"
        ws[f'B{row}'] = "Mã TK"
        ws[f'C{row}'] = "Diễn giải"
        ws[f'D{row}'] = "Số tiền"
        row += 1
        
        for item in data['expense']['details']:
            ws[f'A{row}'] = ""
            ws[f'B{row}'] = item['code']
            ws[f'C{row}'] = item['name']
            ws[f'D{row}'] = item['amount']
            row += 1
        
        ws[f'C{row}'] = "Tổng chi phí"
        ws[f'D{row}'] = data['expense']['total']
        row += 2
        
        ws[f'C{row}'] = "LỢI NHUẬN SAU THUẾ"
        ws[f'D{row}'] = data['net_profit']
        
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 25
        
        wb.save(output_path)
        return output_path
    
    def print_report(self, report_type, params):
        """In báo cáo ra console (kiểm tra nhanh)"""
        if report_type == "financial_position":
            data = self.statement_of_financial_position(**params)
            print("\n" + "="*60)
            print(f"{data['report_name']}")
            print(f"Ngày: {data['report_date']}")
            print("="*60)
            for group_name, group_info in data['groups'].items():
                print(f"\n{group_name}: {group_info['total']:,.0f} VND")
                for acc in group_info['accounts'][:5]:  # Chỉ hiển thị 5 đầu
                    if acc['balance'] != 0:
                        print(f"  {acc['code']} - {acc['name']}: {acc['balance']:,.0f}")
                if len(group_info['accounts']) > 5:
                    print(f"  ... và {len(group_info['accounts'])-5} tài khoản khác")
            print(f"\nTỔNG TÀI SẢN: {data['total_assets']:,.0f} VND")
            print(f"TỔNG NGUỒN VỐN: {data['total_liabilities_equity']:,.0f} VND")
        
        elif report_type == "income":
            data = self.income_statement(**params)
            print("\n" + "="*60)
            print(f"{data['report_name']}")
            print(f"Từ {data['from_date']} đến {data['to_date']}")
            print("="*60)
            print(f"\nDOANH THU: {data['revenue']['total']:,.0f} VND")
            print(f"CHI PHÍ: {data['expense']['total']:,.0f} VND")
            print(f"\nLỢI NHUẬN: {data['net_profit']:,.0f} VND")
        
        return data


# ========== KIỂM TRA NHANH ==========
if __name__ == "__main__":
    print("=== KIỂM TRA MODULE BÁO CÁO TÀI CHÍNH ===")
    fs = FinancialStatements()
    
    # Kiểm tra kết nối database
    if fs.db_path and os.path.exists(fs.db_path):
        print(f"✅ Database: {fs.db_path}")
    else:
        print(f"⚠️ Không tìm thấy database tại: {fs.db_path}")
        print("   Module vẫn hoạt động nhưng sẽ không có dữ liệu")
    
    # In thử báo cáo nếu có dữ liệu
    print("\n--- THỬ IN BÁO CÁO ---")
    fs.print_report("financial_position", {"to_date": datetime.now().strftime("%Y-%m-%d")})