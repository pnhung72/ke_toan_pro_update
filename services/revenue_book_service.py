from models.database import Database
from models.transaction import Transaction
from models.invoice import Invoice
from datetime import datetime
from utils.logger import get_logger
from config_data_source import get_data_source

logger = get_logger(__name__)

class RevenueBookService:
    """Xử lý nghiệp vụ cho mẫu S1a-HKD (Sổ doanh thu bán hàng hóa, dịch vụ)"""
    
    @staticmethod
    def get_revenue_data(ma_so_thue, ky_ke_khai, nam):
        """
        Lấy dữ liệu doanh thu để xuất sổ S1a-HKD
        ky_ke_khai: 'quy1', 'quy2', 'quy3', 'quy4', 'nam'
        nam: năm cần lấy dữ liệu
        """
        from datetime import datetime
        
        # Hàm chuyển đổi chuỗi ngày dd/mm/yyyy thành datetime để so sánh chính xác
        def parse_date(date_str):
            try:
                d, m, y = date_str.split('/')
                return datetime(int(y), int(m), int(d))
            except Exception:
                return None
        
        def is_date_in_range(date_str, start_dt, end_dt):
            """Kiểm tra ngày có nằm trong khoảng không"""
            date_dt = parse_date(date_str)
            if date_dt:
                return start_dt <= date_dt <= end_dt
            return False
        
        # Xác định khoảng thời gian (dùng datetime object)
        if ky_ke_khai == 'quy1':
            start_dt = datetime(nam, 1, 1)
            end_dt = datetime(nam, 3, 31)
            start_date_str = f"01/01/{nam}"
            end_date_str = f"31/03/{nam}"
            ten_ky = f"Quý 1/{nam}"
        elif ky_ke_khai == 'quy2':
            start_dt = datetime(nam, 4, 1)
            end_dt = datetime(nam, 6, 30)
            start_date_str = f"01/04/{nam}"
            end_date_str = f"30/06/{nam}"
            ten_ky = f"Quý 2/{nam}"
        elif ky_ke_khai == 'quy3':
            start_dt = datetime(nam, 7, 1)
            end_dt = datetime(nam, 9, 30)
            start_date_str = f"01/07/{nam}"
            end_date_str = f"30/09/{nam}"
            ten_ky = f"Quý 3/{nam}"
        elif ky_ke_khai == 'quy4':
            start_dt = datetime(nam, 10, 1)
            end_dt = datetime(nam, 12, 31)
            start_date_str = f"01/10/{nam}"
            end_date_str = f"31/12/{nam}"
            ten_ky = f"Quý 4/{nam}"
        else:  # nam
            start_dt = datetime(nam, 1, 1)
            end_dt = datetime(nam, 12, 31)
            start_date_str = f"01/01/{nam}"
            end_date_str = f"31/12/{nam}"
            ten_ky = f"Năm {nam}"
        
        # Lấy tất cả dữ liệu từ transactions (thu từ bán hàng)
        with Database.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT date, description, amount, category, id
                FROM transactions
                WHERE type = 'Thu'
                ORDER BY date
            ''')
            all_transactions = cursor.fetchall()
        
        # Lọc transactions theo ngày thực tế (dùng datetime để so sánh chính xác)
        transactions = []
        for t in all_transactions:
            if is_date_in_range(t['date'], start_dt, end_dt):
                transactions.append({
                    'date': t['date'],
                    'description': t['description'],
                    'amount': t['amount'],
                    'category': t['category'],
                    'id': t['id']
                })
        
        # Lấy dữ liệu từ invoices (hóa đơn) và lọc theo ngày
        all_invoices = Invoice.get_all()
        invoice_data = []
        for inv in all_invoices:
            try:
                d, m, y = inv['created_date'].split('/')
                inv_date = f"{int(d):02d}/{int(m):02d}/{y}"
                if is_date_in_range(inv_date, start_dt, end_dt):
                    invoice_data.append({
                        'date': inv_date,
                        'description': f"Hóa đơn {inv['id']} - {inv['buyer_name']}",
                        'amount': inv['total_payment'],
                        'category': 'Thu hóa đơn',
                        'invoice_id': inv['id']
                    })
            except Exception:
                continue
        
        # === LOGIC PHÁT HIỆN TRÙNG LẶP (giữ nguyên tính năng hữu ích) ===
        # Tạo dictionary để lưu dấu vết giao dịch theo ngày và số tiền
        transaction_map = {}
        for t in transactions:
            key = f"{t['date']}_{t['amount']}"
            if key not in transaction_map:
                transaction_map[key] = []
            transaction_map[key].append({
                'type': 'transaction',
                'date': t['date'],
                'description': t['description'],
                'amount': t['amount'],
                'category': t['category'],
                'id': t['id']
            })
        
        # Phát hiện hóa đơn trùng với giao dịch
        invoice_map = {}
        for inv in invoice_data:
            key = f"{inv['date']}_{inv['amount']}"
            if key not in invoice_map:
                invoice_map[key] = []
            invoice_map[key].append(inv)
        
        # Tìm các key xuất hiện ở cả hai map (trùng lặp)
        duplicate_keys = set(transaction_map.keys()) & set(invoice_map.keys())
        
        # === XÂY DỰNG DANH SÁCH DOANH THU VỚI CẢNH BÁO ===
        all_revenues = []
        duplicate_count = 0
        duplicate_amount = 0
        
        # Thêm giao dịch (đánh dấu nếu trùng)
        for t in transactions:
            key = f"{t['date']}_{t['amount']}"
            is_duplicate = key in duplicate_keys
            
            all_revenues.append({
                'date': t['date'],
                'description': t['description'],
                'amount': t['amount'],
                'category': t['category'],
                'source': 'giao dịch',
                'is_duplicate': is_duplicate
            })
            if is_duplicate:
                duplicate_count += 1
                duplicate_amount += t['amount']
        
        # Thêm hóa đơn (chỉ thêm nếu không trùng, hoặc thêm với ghi chú)
        for inv in invoice_data:
            key = f"{inv['date']}_{inv['amount']}"
            is_duplicate = key in duplicate_keys
            
            if not is_duplicate:
                # Không trùng: thêm bình thường
                all_revenues.append({
                    'date': inv['date'],
                    'description': inv['description'],
                    'amount': inv['amount'],
                    'category': inv['category'],
                    'source': 'hóa đơn',
                    'is_duplicate': False
                })
            else:
                # Có trùng: thêm nhưng đánh dấu và không cộng vào tổng (tránh trùng)
                all_revenues.append({
                    'date': inv['date'],
                    'description': f"{inv['description']} (⚠️ TRÙNG VỚI GIAO DỊCH)",
                    'amount': 0,  # Không cộng để tránh trùng
                    'category': inv['category'],
                    'source': 'hóa đơn (trùng)',
                    'is_duplicate': True,
                    'original_amount': inv['amount']
                })
        
        # Sắp xếp theo ngày (dùng hàm sort với key là datetime để đúng thứ tự)
        def sort_by_date(item):
            d, m, y = item['date'].split('/')
            return (int(y), int(m), int(d))
        
        all_revenues.sort(key=sort_by_date)
        
        # Tính tổng (chỉ tính các mục không trùng và có amount > 0)
        total_revenue = sum(r['amount'] for r in all_revenues if r['amount'] > 0)
        
        # Đếm số lượng giao dịch thực tế
        actual_record_count = len([r for r in all_revenues if r['amount'] > 0])
        
        return {
            'ky_ke_khai': ten_ky,
            'start_date': start_date_str,
            'end_date': end_date_str,
            'revenues': all_revenues,
            'total_revenue': total_revenue,
            'record_count': actual_record_count,
            'raw_record_count': len(all_revenues),
            'duplicate_count': duplicate_count,
            'duplicate_amount': duplicate_amount,
            'has_duplicate': duplicate_count > 0,
            'warning': f"⚠️ Phát hiện {duplicate_count} giao dịch trùng lặp (tổng {duplicate_amount:,.0f} VNĐ). Đã tự động loại bỏ để tránh sai số." if duplicate_count > 0 else None
        }
    
    @staticmethod
    def generate_html_s1a(ma_so_thue, ky_ke_khai, nam, business_info):
        """Tạo HTML cho mẫu S1a-HKD (Sổ doanh thu) - Bảo toàn tính năng"""
        data = RevenueBookService.get_revenue_data(ma_so_thue, ky_ke_khai, nam)
        
        # Xử lý tên hộ kinh doanh để tránh lỗi hiển thị
        ten_ho = business_info.get('ten_ho_kinh_doanh', '')
        dia_chi = business_info.get('dia_chi', '')
        ma_so = business_info.get('ma_so_thue', '')
        ky_ke_khai = data['ky_ke_khai']
        
        html = f"""<!DOCTYPE html>
    <html lang="vi">
    <head>
        <meta charset="UTF-8">
        <title>So doanh thu S1a-HKD - {ten_ho}</title>
        <style>
            body {{ 
                font-family: 'Times New Roman', Arial, sans-serif; 
                margin: 30px; 
                font-size: 13px;
                line-height: 1.4;
            }}
            @media print {{
                body {{ margin: 0; padding: 20px; }}
                .no-print {{ display: none; }}
                .page-break {{ page-break-before: always; }}
            }}
            .header {{ text-align: center; margin-bottom: 25px; }}
            .title {{ font-size: 18px; font-weight: bold; }}
            .subtitle {{ font-size: 13px; color: #333; }}
            .info-table {{ width: 100%; border-collapse: collapse; margin-top: 20px; margin-bottom: 20px; }}
            .info-table td, .info-table th {{ border: 1px solid #000; padding: 8px; vertical-align: top; }}
            .info-table td:first-child {{ width: 25%; background-color: #f5f5f5; font-weight: bold; }}
            .data-table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            .data-table th, .data-table td {{ border: 1px solid #000; padding: 8px; }}
            .data-table th {{ background-color: #e0e0e0; font-weight: bold; text-align: center; }}
            .data-table td:nth-child(1) {{ text-align: center; width: 15%; }}
            .data-table td:nth-child(2) {{ text-align: left; width: 65%; }}
            .data-table td:nth-child(3) {{ text-align: right; width: 20%; }}
            .total-row {{ font-weight: bold; background-color: #f0f0f0; }}
            .duplicate-row {{ background-color: #ffeeee; color: #999; text-decoration: line-through; }}
            .warning {{ background-color: #ffcccc; color: red; padding: 10px; margin: 15px 0; border: 1px solid red; border-radius: 5px; }}
            .info {{ background-color: #e8f4fd; color: #2196F3; padding: 10px; margin: 15px 0; border: 1px solid #2196F3; border-radius: 5px; }}
            .signature {{ margin-top: 50px; text-align: right; }}
            .footer {{ font-size: 11px; margin-top: 30px; text-align: center; border-top: 1px solid #ccc; padding-top: 10px; }}
            .stt-col {{ text-align: center; width: 5%; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="title">SỔ DOANH THU BÁN HÀNG HÓA, DỊCH VỤ</div>
            <div class="subtitle">(Mẫu số S1a-HKD ban hành kèm theo Thông tư 152/2025/TT-BTC)</div>
        </div>
        
        <table class="info-table">
            <tr><td style="width: 25%;">Hộ kinh doanh:</td><td colspan="3">{ten_ho}</td></tr>
            <tr><td>Mã số thuế:</td><td colspan="3">{ma_so}</td></tr>
            <tr><td>Địa chỉ:</td><td colspan="3">{dia_chi}</td></tr>
            <tr><td>Kỳ kê khai:</td><td colspan="3">{ky_ke_khai}</td></tr>
            <tr><td>Địa điểm kinh doanh:</td><td colspan="3">{dia_chi}</td></tr>
        </table>
    """
        
        if data.get('has_duplicate', False):
            html += f"""
        <div class="warning">
            ⚠️ {data.get('warning', 'Phát hiện giao dịch trùng lặp!')}
        </div>
    """
        
        # Thêm thông tin nguồn dữ liệu (tính năng hữu ích)
        if data.get('source_name'):
            html += f"""
        <div class="info">
            📊 Nguồn dữ liệu: {data.get('source_name', 'Chưa xác định')}
        </div>
    """
        
        html += """
        <table class="data-table">
            <thead>
                <tr>
                    <th class="stt-col">STT</th>
                    <th>Ngày, tháng ghi sổ</th>
                    <th>Diễn giải</th>
                    <th>Số tiền (VNĐ)</th>
                </tr>
            </thead>
            <tbody>
    """
        
        stt = 1
        for rev in data['revenues']:
            # Xử lý ký tự đặc biệt trong description
            description = rev['description']
            description = description.replace('&', '&amp;')
            description = description.replace('<', '&lt;')
            description = description.replace('>', '&gt;')
            
            # Xác định class cho dòng
            row_class = ''
            amount_display = ''
            
            if rev.get('is_duplicate') and rev.get('amount', 0) == 0:
                row_class = 'duplicate-row'
                original_amount = rev.get('original_amount', 0)
                amount_display = f"<span style='color: #999;'>(Đã loại trùng) {original_amount:,.0f}</span>"
            else:
                amount_display = f"{rev['amount']:,.0f}"
            
            html += f"""
                <tr class="{row_class}">
                    <td style="text-align: center;">{stt}</td>
                    <td style="text-align: center;">{rev['date']}</td>
                    <td>{description}</td>
                    <td style="text-align: right;">{amount_display}</td>
                </tr>
    """
            stt += 1
        
        html += f"""
            </tbody>
            <tfoot>
                <tr class="total-row">
                    <td colspan="3" style="text-align: right; font-weight: bold;">Tổng cộng ({data['record_count']} giao dịch):</td>
                    <td style="text-align: right; font-weight: bold;">{data['total_revenue']:,.0f}</td>
                </tr>
            </tfoot>
        </table>
        
        <div class="signature">
            <div>Ngày ... tháng ... năm ...</div>
            <div style="margin-top: 40px;"><strong>NGƯỜI ĐẠI DIỆN HỘ KINH DOANH</strong></div>
            <div>(Ký, ghi rõ họ tên, đóng dấu nếu có)</div>
        </div>
        
        <div class="footer">
            <hr>
            <div>📌 Lưu ý: Sổ này dùng để ghi chép doanh thu bán hàng hóa, dịch vụ theo quy định tại Thông tư 152/2025/TT-BTC.</div>
            <div>📌 Hộ kinh doanh có trách nhiệm lưu giữ sổ và xuất trình khi cơ quan thuế yêu cầu.</div>
            <div>📌 Thời hạn lưu giữ: 10 năm kể từ ngày kết thúc năm tài chính.</div>
        </div>
        
        <div class="no-print" style="text-align: center; margin-top: 20px;">
            <button onclick="window.print();" style="padding: 8px 20px; font-size: 14px; cursor: pointer;">🖨️ In sổ doanh thu</button>
            <button onclick="window.close();" style="padding: 8px 20px; font-size: 14px; cursor: pointer;">❌ Đóng</button>
        </div>
    </body>
    </html>
    """
        return html
    
    @staticmethod
    def generate_excel_s1a(ma_so_thue, ky_ke_khai, nam, business_info, output_path):
        """Xuất Excel cho mẫu S1a-HKD"""
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        
        data = RevenueBookService.get_revenue_data(ma_so_thue, ky_ke_khai, nam)
        
        # Tạo DataFrame
        df_data = []
        for rev in data['revenues']:
            df_data.append({
                'Ngày ghi sổ': rev['date'],
                'Diễn giải': rev['description'],
                'Số tiền (VNĐ)': rev['amount']
            })
        
        df = pd.DataFrame(df_data)
        
        # Thêm dòng tổng
        total_row = pd.DataFrame({
            'Ngày ghi sổ': ['Tổng cộng:'],
            'Diễn giải': [''],
            'Số tiền (VNĐ)': [data['total_revenue']]
        })
        df = pd.concat([df, total_row], ignore_index=True)
        
        # Xuất Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='So doanh thu S1a-HKD', index=False)
            
            # Format
            workbook = writer.book
            worksheet = writer.sheets['So doanh thu S1a-HKD']
            
            # Định dạng tiền tệ
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=3):
                for cell in row:
                    cell.number_format = '#,##0'
            
            # Điều chỉnh độ rộng cột
            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 50
            worksheet.column_dimensions['C'].width = 20
        
        return output_path