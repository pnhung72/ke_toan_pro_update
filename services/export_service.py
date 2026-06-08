# -*- coding: utf-8 -*-
"""
ExportService - Xuat bao cao PDF, Excel, XML
"""

import os
from utils.logger import get_logger
logger = get_logger(__name__)
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class ExportService:
    """Dich vu xuat bao cao"""
    
    def __init__(self, export_dir="exports"):
        self.export_dir = export_dir
        os.makedirs(export_dir, exist_ok=True)
    
    def export_to_excel(self, data, filename, sheet_name="Sheet1"):
        """Xuat du lieu ra file Excel"""
        try:
            filepath = os.path.join(self.export_dir, f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            df = pd.DataFrame(data)
            df.to_excel(filepath, sheet_name=sheet_name, index=False)
            logger.info(f"Đã xuất Excel: {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"Lỗi xuất Excel: {e}")
            return None
    
    def export_to_excel_styled(self, headers, data, filename, title=None):
        """Xuat Excel co dinh dang"""
        filepath = os.path.join(self.export_dir, f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        
        wb = Workbook()
        ws = wb.active
        
        # Tieu de
        if title:
            ws.merge_cells('A1:{}1'.format(chr(64 + len(headers))))
            cell = ws['A1']
            cell.value = title
            cell.font = Font(size=14, bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Header
        row_start = 3 if title else 1
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_start, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_idx, row_data in enumerate(data, row_start + 1):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(filepath)
        return filepath
    
    def export_to_xml(self, data, root_name, filename):
        """Xuat du lieu ra XML"""
        import xml.etree.ElementTree as ET
        
        filepath = os.path.join(self.export_dir, f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xml")
        
        root = ET.Element(root_name)
        
        for item in data:
            item_elem = ET.SubElement(root, "item")
            for key, value in item.items():
                child = ET.SubElement(item_elem, key)
                child.text = str(value)
        
        tree = ET.ElementTree(root)
        tree.write(filepath, encoding='utf-8', xml_declaration=True)
        
        return filepath
    
    def export_revenue_report(self, revenue_data, year, month):
        """Xuat bao cao doanh thu"""
        headers = ["Ngay", "Doanh thu", "Thue VAT", "Giam gia", "Thuc thu"]
        
        data = []
        for day, values in revenue_data.items():
            data.append([
                day,
                f"{values['revenue']:,.0f}",
                f"{values.get('vat', 0):,.0f}",
                f"{values.get('discount', 0):,.0f}",
                f"{values.get('actual', values['revenue']):,.0f}"
            ])
        
        title = f"BAO CAO DOANH THU - Thang {month}/{year}"
        return self.export_to_excel_styled(headers, data, f"doanh_thu_{year}_{month}", title)
    
    def export_debt_report(self, debt_data):
        """Xuat bao cao cong no"""
        headers = ["Khach hang", "Tong cong no", "Da tra", "Con lai", "Han tra", "Trang thai"]
        
        data = []
        for customer in debt_data:
            data.append([
                customer.get("name", ""),
                f"{customer.get('total_debt', 0):,.0f}",
                f"{customer.get('paid', 0):,.0f}",
                f"{customer.get('remaining', 0):,.0f}",
                customer.get("due_date", ""),
                "Qua han" if customer.get("is_overdue", False) else "Con han"
            ])
        
        title = "BAO CAO CONG NO"
        return self.export_to_excel_styled(headers, data, "cong_no", title)
    
    def export_invoice_xml(self, invoice_data):
        """Xuat hoa don ra XML theo chuan TCT"""
        root_name = "HoaDon"
        
        data = [{
            "SoHoaDon": invoice_data.get("id", ""),
            "NgayLap": invoice_data.get("date", ""),
            "KhachHang": invoice_data.get("customer", ""),
            "TongTien": invoice_data.get("total", 0),
            "ThueVAT": invoice_data.get("vat", 0),
            "ThanhTien": invoice_data.get("total", 0) + invoice_data.get("vat", 0)
        }]
        
        return self.export_to_xml(data, root_name, f"hoa_don_{invoice_data.get('id', '')}")